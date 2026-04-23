"""Чтение и запись Excel-файлов для SOP Validator.

Вход: лист "Questions" с колонками sop_id | question | question_type | reference_answer
      (опционально: options — варианты ответов для multiple).

Выход — одна из двух схем:

detailed=False (по умолчанию, «бизнес-вид»):
    - per-arch листы: question | reference_answer | llm_answer | verdict
    - Metrics: architecture | total | correct | accuracy

detailed=True («DS-вид»):
    - per-arch листы + Metrics с Wilson CI, токенами, latency, per-type accuracy
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import (
    ArchitectureName,
    InputQuestion,
    QuestionRunResult,
)


REQUIRED_COLUMNS = ["sop_id", "question", "question_type", "reference_answer"]
OPTIONAL_COLUMNS = ["options"]
INPUT_SHEET_NAME = "Questions"

SHEET_TITLE = {
    ArchitectureName.ZERO_SHOT: "ZeroShot",
    ArchitectureName.COT: "CoT",
    ArchitectureName.SELF_CONSISTENCY: "SelfConsistency",
}

BASIC_COLUMNS = ["question", "reference_answer", "llm_answer", "verdict"]


def read_questions(xlsx_path: str | Path) -> list[InputQuestion]:
    """Прочитать и валидировать входной Excel."""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Входной Excel не найден: {xlsx_path}")

    try:
        df = pd.read_excel(xlsx_path, sheet_name=INPUT_SHEET_NAME, dtype=str)
    except ValueError:
        logger.warning(
            f"Лист '{INPUT_SHEET_NAME}' не найден в {xlsx_path.name}, использую первый лист"
        )
        df = pd.read_excel(xlsx_path, sheet_name=0, dtype=str)

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"В входном Excel отсутствуют обязательные колонки: {missing}. "
            f"Ожидается: {REQUIRED_COLUMNS}"
        )

    df = df.fillna("")
    questions: list[InputQuestion] = []
    skipped = 0
    for idx, row in df.iterrows():
        try:
            q = InputQuestion(
                sop_id=str(row["sop_id"]).strip(),
                question=str(row["question"]),
                question_type=str(row["question_type"]).strip().lower(),
                reference_answer=str(row["reference_answer"]),
                options=str(row.get("options", "")).strip(),
            )
            questions.append(q)
        except Exception as exc:
            skipped += 1
            logger.warning(f"Пропускаю строку {idx + 2} (Excel-нумерация): {exc}")

    logger.info(f"Загружено вопросов: {len(questions)} (пропущено: {skipped})")
    if not questions:
        raise ValueError("После валидации не осталось ни одного корректного вопроса")
    return questions


def _format_verdict(result: QuestionRunResult, arch: ArchitectureName) -> str:
    """Для закрытых вопросов — «правильно»/«неправильно». Для open — с комментарием судьи."""
    verdict = result.verdicts.get(arch)
    if verdict is None or verdict.judge_failed or verdict.score is None:
        return "не определено"
    label = "правильно" if verdict.score == 1 else "неправильно"
    if result.input.is_closed:
        return label
    comment = (verdict.comment or "").strip()
    return f"{label} — {comment}" if comment and comment != "(без комментария)" else label


def _arch_sheet_df(results: list[QuestionRunResult], arch: ArchitectureName) -> pd.DataFrame:
    rows = []
    for r in results:
        arch_res = r.results.get(arch)
        rows.append({
            "question": r.input.question,
            "reference_answer": r.input.reference_answer,
            "llm_answer": arch_res.answer if arch_res else "",
            "verdict": _format_verdict(r, arch),
        })
    return pd.DataFrame(rows, columns=BASIC_COLUMNS)


def write_results(
    output_path: str | Path,
    results: list[QuestionRunResult],
    enabled_archs: list[ArchitectureName],
    metrics_df: pd.DataFrame | None = None,
) -> None:
    """Записать результаты в Excel: лист на архитектуру + опциональный Metrics."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for arch in enabled_archs:
            _arch_sheet_df(results, arch).to_excel(
                writer, sheet_name=SHEET_TITLE[arch], index=False
            )
        if metrics_df is not None and not metrics_df.empty:
            metrics_df.to_excel(writer, sheet_name="Metrics", index=False)

    _apply_formatting(output_path)
    logger.info(f"Результаты записаны: {output_path}")


def _apply_formatting(path: Path) -> None:
    wb = load_workbook(path)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    body_font = Font(name="Arial", size=10)

    col_widths = {"question": 60, "reference_answer": 40, "llm_answer": 50, "verdict": 25}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1:
            continue
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col_idx).value or "").lower()
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = col_widths.get(header, 20)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
                if cell.font is None or cell.font.name is None:
                    cell.font = body_font
        ws.freeze_panes = "A2"
    wb.save(path)


class CheckpointWriter:
    """Периодическое сохранение промежуточных результатов."""

    def __init__(self, output_path: str | Path, enabled_archs: list[ArchitectureName], every_n: int) -> None:
        self.output_path = Path(output_path)
        self.enabled_archs = enabled_archs
        self.every_n = max(1, every_n)
        self._buffer: list[QuestionRunResult] = []

    def add(self, result: QuestionRunResult) -> None:
        self._buffer.append(result)
        if len(self._buffer) % self.every_n == 0:
            self.flush()

    def flush(self) -> None:
        if not self._buffer:
            return
        checkpoint_path = self.output_path.with_suffix(".checkpoint.xlsx")
        write_results(checkpoint_path, self._buffer, self.enabled_archs)
        logger.debug(f"Checkpoint: {len(self._buffer)} вопросов → {checkpoint_path.name}")
